import cv2
import numpy as np
import face_recognition
import os

from datetime import datetime, timedelta
from time import sleep
from openpyxl.styles import Font
from openpyxl import load_workbook
import threading
from threading import Lock
import sys
import os
# Dynamically add the utils_folder to the Python path
sys.path.append(os.path.abspath('../'))
from Excel_Format import (get_top_left_cell, search_keyword_ranges, find_end_limit, update_student_info, get_current_session, get_or_create_today_sheet)
print(os.getcwd())
# Load the workbook
file_path = 'Book1.xlsx'
  # Path to your attendance workbook
workbook = load_workbook(file_path)
target_sheet = get_or_create_today_sheet(workbook)

# Global dictionary to store student status and timestamps
student_status = {}

# Absence threshold and timers
absence_threshold = timedelta(seconds=5)  # Set your desired threshold for absence
extended_absence_timer = timedelta(seconds=10)  # Set the time after which a student is considered permanently absent
grace_period = timedelta(seconds=1)  # Set the grace period for students who have been marked absent


# print("Mark Attendance by : \n 1. Name \n 2. Roll No. \n")
# No = int(input("Mark Attendance by : \n 1. Name \n 2. Roll No. \n"))
No=1
# Path to training images
path1 = '../Training_images/Name'
path2 = '../Training_images/Roll No.'
images = []
classNames = []



# Load images and class names
if No == 1:
    myList = os.listdir(path1)
    print("Training images found:", myList)

    for cl in myList:
        curImg = cv2.imread(f'{path1}/{cl}')
        images.append(curImg)
        classNames.append(os.path.splitext(cl)[0])
elif No == 2:
    myList = os.listdir(path2)
    print("Training images found:", myList)

    for cl in myList:
        curImg = cv2.imread(f'{path2}/{cl}')
        images.append(curImg)
        classNames.append(os.path.splitext(cl)[0])
else:
    print("Invalid input. Please enter 1 or 2.")
    exit()

print("Class names:", classNames)

def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encodings = face_recognition.face_encodings(img)
        if encodings:
            encodeList.append(encodings[0])
    return encodeList

def search_name_and_mark_attendance(sheet, name, current_session_start_row, current_session_end_row):
    """Search for the name in the given row range, mark attendance for the given name as 'Present',
       and mark the remaining students as 'Absent', skipping headers."""
    
    name = name.upper()
    
    # Adjust offsets based on the new layout
    name_column_offset = 3  # Actual name column (after Roll NO and PRN no)
    status_column_offset = 1  # Status column offset
    last_seen_column_offset = 2  # 'Last Seen' column offset
    absence_timer_column_offset = 3  # 'Start Absence Timer' column offset

    current_time = datetime.now()  # Get the current time
    current_time_str = current_time.strftime('%H:%M:%S')  # Current time string for the sheet
    name_found = False  # Track if the student has been found

    # Step 1: Mark the given student as 'Present' in the current session
    for row_idx, row in enumerate(sheet.iter_rows(min_row=current_session_start_row, max_row=current_session_end_row), start=current_session_start_row):
        if row_idx == current_session_start_row or isinstance(row[name_column_offset].value, str) and row[name_column_offset].value.upper() == "name":
            continue  # Don't process the header row
        
        for cell in row:
            if isinstance(cell.value, str) and cell.value.upper() == name:
                name_found = True  # Name found
                
                # Mark 'Present'
                status_cell = get_top_left_cell(sheet, sheet.cell(row=cell.row, column=cell.column + status_column_offset))
                if status_cell.value != "Present":
                    status_cell.value = "Present"
                    status_cell.font = Font(color="006400", bold=True)  # Green font for Present
                    print(f"Marked {name} as Present.")

                # Update 'Last Seen'
                last_seen_cell = get_top_left_cell(sheet, sheet.cell(row=cell.row, column=cell.column + last_seen_column_offset))
                last_seen_cell.value = current_time_str
                last_seen_cell.font = Font(color="006400", bold=True)

                # Update 'Start Absence Timer'
                absence_timer_cell = get_top_left_cell(sheet, sheet.cell(row=cell.row, column=cell.column + absence_timer_column_offset))
                absence_timer_cell.value = current_time_str
                absence_timer_cell.font = Font(color="006400", bold=True)

                # Update global student status
                student_status[name] = (current_time, 'Present', None)  # Add None for Start Absence Timer
                print(f"Updated status for {name}: Present at {current_time_str}")
                # continuously_check_absence(target_sheet)
                break

    # Step 2: Mark remaining students as 'Absent'
    for row_idx, row in enumerate(sheet.iter_rows(min_row=current_session_start_row, max_row=current_session_end_row), start=current_session_start_row):
        if row_idx == current_session_start_row or isinstance(row[name_column_offset].value, str) and row[name_column_offset].value.upper() == "name":
            continue  # Don't process the header row
        
        name_cell = row[name_column_offset]
        if isinstance(name_cell.value, str):
            status_cell = get_top_left_cell(sheet, sheet.cell(row=name_cell.row, column=name_cell.column + status_column_offset))
            student_name = name_cell.value.upper()

            if student_name not in student_status:
                student_status[student_name] = (None, 'Absent', None)  # Initialize if not present

            # If the student is not the one just marked as present
            if student_name != name:
                last_seen_time, status, absence_timer_start = student_status[student_name]

                if last_seen_time is None:
                    # Mark as Absent
                    if status_cell.value is None or status_cell.value.strip() == "":
                        status_cell.value = "Absent"
                        status_cell.font = Font(color="FF0000")  # Red font for Absent
                        last_seen_cell = get_top_left_cell(sheet, sheet.cell(row=name_cell.row, column=name_cell.column + last_seen_column_offset))
                        last_seen_cell.value = "N/A"
                        last_seen_cell.font = Font(color="FF0000")
                        absence_timer_cell = get_top_left_cell(sheet, sheet.cell(row=name_cell.row, column=name_cell.column + absence_timer_column_offset))
                        absence_timer_cell.value = "Not started"
                        absence_timer_cell.font = Font(color="FF0000")
                else:
                    time_since_last_seen = current_time - last_seen_time

                    # Mark as Permanently Absent
                    if time_since_last_seen > extended_absence_timer:
                        if status_cell.value is None or status_cell.value.strip() == "":
                            status_cell.value = "Permanently Absent"
                            status_cell.font = Font(color="FFA500")  # Orange font for Permanently Absent
                            print(f"Marked {student_name} as Permanently Absent due to extended absence.")
                    elif time_since_last_seen > absence_threshold:
                        if time_since_last_seen > grace_period:
                            # Mark as Absent
                            if status_cell.value is None or status_cell.value.strip() == "":
                                status_cell.value = "Absent"
                                status_cell.font = Font(color="FF0000")  # Red font for Absent
                                print(f"Marked {student_name} as Absent after exceeding absence threshold and grace period.")

    return name_found  # Return whether the name was found or not

def format_duration(seconds):
    """Format a duration in seconds into a more readable string."""
    if seconds < 60:
        return f"{seconds} seconds"
    elif seconds < 3600:  # Less than an hour
        minutes = seconds // 60
        seconds = seconds % 60
        return f"{minutes} minutes, {seconds} seconds"
    else:  # Greater than or equal to an hour
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        seconds = seconds % 60
        return f"{hours} hours, {minutes} minutes, {seconds} seconds"

# Initialize a thread-safe lock
sheet_lock = Lock()

# Example student status dictionary
student_status = {
    # 'Student Name': (last_seen_time, status, absence_timer_start)
}

def continuously_check_absence(sheet, keyword, interval=5, absence_threshold=timedelta(seconds=5), 
                               grace_period=timedelta(seconds=7), extended_absence_timer=timedelta(seconds=10)):
    """
    Continuously checks the absence status of students based on their last seen time.
    Updates the Excel sheet with the status and handles transitions between Present, Temporary Absent, and Permanently Absent.
    """
    # Determine start and end rows based on the session keyword
    start_rows = search_keyword_ranges(sheet, keyword)
    if not start_rows:
        print(f"No rows found for keyword '{keyword}'.")
        return

    # Get the first found row for the session, skipping the header
    start_row = start_rows[0] + 1
    end_row = find_end_limit(sheet, start_row)  # Determine end row based on the start row

    # Offsets for Excel columns
    status_column_offset = 1
    last_seen_column_offset = 2
    absence_timer_start_column_offset = 3

    while True:
        current_time = datetime.now()
        print(f"Checking absence status at {current_time.strftime('%H:%M:%S')}")

        for student_name, (last_seen_time, status, absence_timer_start) in student_status.items():
            if last_seen_time is None:
                # Skip students without a last_seen_time
                continue

            time_since_last_seen = current_time - last_seen_time
            print(f"Time since last seen for {student_name}: {time_since_last_seen}")

            if status == "Present":
                if time_since_last_seen > absence_threshold:
                    if absence_timer_start is None:
                        absence_timer_start = current_time
                        student_status[student_name] = (last_seen_time, status, absence_timer_start)
                        with sheet_lock:
                            update_student_info(sheet, student_name, last_seen_time, current_time, 
                                                start_row, end_row, status_column_offset, last_seen_column_offset, 
                                                absence_timer_start_column_offset, status="Absent Timer Started")
                        print(f"Started absence timer for {student_name} at {current_time.strftime('%H:%M:%S')}.")
                    else:
                        # Check how long the timer has been running
                        time_since_timer_started = current_time - absence_timer_start
                        time_in_seconds = int(time_since_timer_started.total_seconds())
                        print(f"Absence timer for {student_name}: {time_in_seconds} seconds")

                        if time_since_timer_started > grace_period:
                            if time_since_timer_started > extended_absence_timer:
                                with sheet_lock:
                                    update_student_info(sheet, student_name, last_seen_time, current_time, 
                                                        start_row, end_row, status_column_offset, last_seen_column_offset, 
                                                        absence_timer_start_column_offset, status="Temporary Absent", 
                                                        update_absence_timer=False, time_in_seconds=time_in_seconds)
                                student_status[student_name] = (last_seen_time, "Temporary Absent", absence_timer_start)
                                print(f"Status for {student_name} updated to Temporary Absent.")

                            if time_since_timer_started > (grace_period + extended_absence_timer):
                                with sheet_lock:
                                    update_student_info(sheet, student_name, last_seen_time, current_time, 
                                                        start_row, end_row, status_column_offset, last_seen_column_offset, 
                                                        absence_timer_start_column_offset, status="Permanently Absent", 
                                                        time_in_seconds=time_in_seconds)
                                student_status[student_name] = (last_seen_time, "Permanently Absent", None)
                                print(f"Status for {student_name} updated to Permanently Absent.")

        sleep(interval)

def process_face_recognition_results(face_recognition_results, student_status, current_time, sheet, start_row, end_row):
    """
    Process face recognition results and update the status of students based on recognition and absence timer.
    """
    for student_name, details in student_status.items():
        last_seen_time, status, absence_timer_start = details
        
        if student_name in face_recognition_results:  # Face recognized
            if status in ["Temporary Absent"]:
                print(f"Face recognized again for {student_name}. Resetting status to 'Present'.")
                status = "Present"
            
            # Update last seen time to the current time
            last_seen_time = current_time
            absence_timer_start = None  # Reset absence timer
            print(f"Updated last seen time for {student_name} to {current_time.strftime('%H:%M:%S')}.")
        
        else:  # Face not recognized
            if last_seen_time is None:  # No prior record of being seen
                continue

            time_since_last_seen = current_time - last_seen_time

            if status == "Present" and time_since_last_seen > timedelta(seconds=5):  # Absence threshold
                if absence_timer_start is None:
                    absence_timer_start = current_time  # Start absence timer
                    print(f"Started absence timer for {student_name} at {current_time.strftime('%H:%M:%S')}.")
                else:
                    time_since_timer_started = current_time - absence_timer_start
                    if time_since_timer_started > timedelta(seconds=10):  # Extended absence threshold
                        status = "Permanently Absent"
                        print(f"{student_name} marked as Permanently Absent.")
                    elif time_since_timer_started > timedelta(seconds=7):  # Grace period
                        status = "Temporary Absent"
                        print(f"{student_name} marked as Temporary Absent.")
        
        # Update the student status in memory
        student_status[student_name] = (last_seen_time, status, absence_timer_start)


        
        # Update Excel sheet
        update_student_info(sheet, student_name, last_seen_time, current_time, start_row, end_row, status)

    return student_status

keyword = get_current_session()
# Start the absence checking in a separate thread
# Ensure you define the `current_session_start_row` and `current_session_end_row` based on your context
absence_check_thread = threading.Thread(target=continuously_check_absence, args=(target_sheet, keyword))
absence_check_thread.daemon = True  # Daemonize thread to ensure it exits when the main program exits
absence_check_thread.start()

def draw_wrapped_text_centered(img, text, x, y, rect_width, color, font_scale=1, thickness=1):
        words = text.split(' ')
        lines = []
        current_line = words[0]
        
        for word in words[1:]:
            (w, h), _ = cv2.getTextSize(current_line + ' ' + word, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness)
            if w > rect_width - 10:
                lines.append(current_line)
                current_line = word
            else:
                current_line += ' ' + word
        lines.append(current_line)

        for i, line in enumerate(lines):
            (line_width, line_height), _ = cv2.getTextSize(line, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness)
            line_x = x + (rect_width - line_width) // 2
            y_offset = y + i * (line_height + 5)
            cv2.putText(img, line, (line_x, y_offset), cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 255, 255), thickness)

        return len(lines)


def search_roll_no_and_mark_attendance(sheet, roll_no, current_session_start_row, current_session_end_row):
    """Search for the roll_no in the given row range, mark attendance for the given roll_no as 'Present',
       and mark the remaining students as 'Absent', skipping headers."""
    
    roll_no = roll_no.upper()
    
    # Adjust offsets based on the new layout
    roll_no_column_offset = 3  # Actual roll_no column (after Roll NO and PRN no)
    status_column_offset = 1  # Status column offset
    last_seen_column_offset = 2  # 'Last Seen' column offset
    absence_timer_column_offset = 3  # 'Start Absence Timer' column offset

    current_time = datetime.now()  # Get the current time
    current_time_str = current_time.strftime('%H:%M:%S')  # Current time string for the sheet
    roll_no_found = False  # Track if the student has been found

    # Step 1: Mark the given student as 'Present' in the current session
    for row_idx, row in enumerate(sheet.iter_rows(min_row=current_session_start_row, max_row=current_session_end_row), start=current_session_start_row):
        if row_idx == current_session_start_row or isinstance(row[roll_no_column_offset].value, str) and row[roll_no_column_offset].value.upper() == "roll_no":
            continue  # Don't process the header row
        
        for cell in row:
            if isinstance(cell.value, str) and cell.value.upper() == roll_no:
                roll_no_found = True  # roll_no found
                
                # Mark 'Present'
                status_cell = get_top_left_cell(sheet, sheet.cell(row=cell.row, column=cell.column + status_column_offset))
                if status_cell.value != "Present":
                    status_cell.value = "Present"
                    status_cell.font = Font(color="006400", bold=True)  # Green font for Present
                    print(f"Marked {roll_no} as Present.")

                # Update 'Last Seen'
                last_seen_cell = get_top_left_cell(sheet, sheet.cell(row=cell.row, column=cell.column + last_seen_column_offset))
                last_seen_cell.value = current_time_str
                last_seen_cell.font = Font(color="006400", bold=True)

                # Update 'Start Absence Timer'
                absence_timer_cell = get_top_left_cell(sheet, sheet.cell(row=cell.row, column=cell.column + absence_timer_column_offset))
                absence_timer_cell.value = current_time_str
                absence_timer_cell.font = Font(color="006400", bold=True)

                # Update global student status
                student_status[roll_no] = (current_time, 'Present', None)  # Add None for Start Absence Timer
                print(f"Updated status for {roll_no}: Present at {current_time_str}")
                # continuously_check_absence(target_sheet)
                break

    # Step 2: Mark remaining students as 'Absent'
    for row_idx, row in enumerate(sheet.iter_rows(min_row=current_session_start_row, max_row=current_session_end_row), start=current_session_start_row):
        if row_idx == current_session_start_row or isinstance(row[roll_no_column_offset].value, str) and row[roll_no_column_offset].value.upper() == "roll no":
            continue  # Don't process the header row
        
        roll_no_cell = row[roll_no_column_offset]
        if isinstance(roll_no_cell.value, str):
            status_cell = get_top_left_cell(sheet, sheet.cell(row=roll_no_cell.row, column=roll_no_cell.column + status_column_offset))
            student_roll_no = roll_no_cell.value.upper()

            if student_roll_no not in student_status:
                student_status[student_roll_no] = (None, 'Absent', None)  # Initialize if not present

            # If the student is not the one just marked as present
            if student_roll_no != roll_no:
                last_seen_time, status, absence_timer_start = student_status[student_roll_no]

                if last_seen_time is None:
                    # Mark as Absent
                    if status_cell.value is None or status_cell.value.strip() == "":
                        status_cell.value = "Absent"
                        status_cell.font = Font(color="FF0000")  # Red font for Absent
                        last_seen_cell = get_top_left_cell(sheet, sheet.cell(row=roll_no_cell.row, column=roll_no_cell.column + last_seen_column_offset))
                        last_seen_cell.value = "N/A"
                        last_seen_cell.font = Font(color="FF0000")
                        absence_timer_cell = get_top_left_cell(sheet, sheet.cell(row=roll_no_cell.row, column=roll_no_cell.column + absence_timer_column_offset))
                        absence_timer_cell.value = "Not started"
                        absence_timer_cell.font = Font(color="FF0000")
                else:
                    time_since_last_seen = current_time - last_seen_time

                    # Mark as Permanently Absent
                    if time_since_last_seen > extended_absence_timer:
                        if status_cell.value is None or status_cell.value.strip() == "":
                            status_cell.value = "Permanently Absent"
                            status_cell.font = Font(color="FFA500")  # Orange font for Permanently Absent
                            print(f"Marked {student_roll_no} as Permanently Absent due to extended absence.")
                    elif time_since_last_seen > absence_threshold:
                        if time_since_last_seen > grace_period:
                            # Mark as Absent
                            if status_cell.value is None or status_cell.value.strip() == "":
                                status_cell.value = "Absent"
                                status_cell.font = Font(color="FF0000")  # Red font for Absent
                                print(f"Marked {student_roll_no} as Absent after exceeding absence threshold and grace period.")

    return roll_no_found  # Return whether the roll_no was found or not

# Load encodings for face recognition
encodeListKnown = findEncodings(images)
print('Encoding Complete')
# videopath = r"F:\D29_20241227100000_001.mp4"
cap = cv2.VideoCapture(0)  # Change the index if necessary

# Check if the camera is opened successfully
if not cap.isOpened():
    print("Error: Could not open camera.")
    cap.release()  # Release any resources
else:
    print("Camera opened successfully.")

last_printed_session = None  # To avoid printing session details multiple times
last_recognized_face = {}  # Dictionary to track faces already printed in this session

try:
    while True:
        success, img = cap.read()
        if not success:
            print("Failed to capture frame from webcam. Exiting...")
            break

        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        # Find faces in the current frame
        faceCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

        current_session = get_current_session()

        if current_session:
            # Check if the current session has changed
            if current_session != last_printed_session:
                print(f"Current Session: {current_session}")
                last_printed_session = current_session  # Update the last printed session
                last_recognized_face = {}  # Reset recognized faces for the new session
        else:
            print("No active session at this time.")
            continue  # Skip to the next iteration if no active session

        # Loop through all faces detected in the current frame
        for encodeFace, faceLoc in zip(encodesCurFrame, faceCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDist = face_recognition.face_distance(encodeListKnown, encodeFace)

            # Identify the best match
            matchIndex = np.argmin(faceDist)

            if matches[matchIndex]:
                name = classNames[matchIndex].upper()
                color = (0, 255, 0)
            else:
                name = "UNKNOWN"
                color = (0, 0, 255)

            # Draw a rectangle around the face and display the name
            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4

            font_scale = 0.7
            (text_width, text_height), _ = cv2.getTextSize(name, cv2.FONT_HERSHEY_SIMPLEX, font_scale, 2)
            rect_width = x2 - x1
            rect_height = text_height * 2 + 5

            cv2.rectangle(img, (x1, y1), (x2, y2), color, 2)
            cv2.rectangle(img, (x1, y2 - rect_height), (x2, y2), color, cv2.FILLED)

            draw_wrapped_text_centered(img, name, x1, y2 - rect_height + text_height, rect_width, color, font_scale=font_scale, thickness=2)

            # Only print and mark attendance if it's a new face for this session
            if name != "UNKNOWN" and name not in last_recognized_face:
                print(f"Best match: {name}, Face distances: {faceDist}")
                last_recognized_face[name] = True  # Mark this face as already printed for this session

                # Check the session and mark attendance
                session_rows = search_keyword_ranges(target_sheet, current_session)
                print(f"Session '{current_session}' found at rows: {session_rows}")

                if session_rows:
                    start_row = session_rows[0]
                    end_row = find_end_limit(target_sheet, start_row, keyword_column=1)
                    print(f"End row for session starting at row {start_row}: {end_row}")
                    print(f"Searching for name '{name}' between rows {start_row} and {end_row}")
                    
                    if No == 1:
                        name_found = search_name_and_mark_attendance(target_sheet, name, start_row, end_row)
                        Found = name_found
                    elif No == 2:
                        roll_no_found = search_roll_no_and_mark_attendance(target_sheet, name, start_row, end_row)
                        Found = roll_no_found
                    else:
                        print("Invalid input. Please enter 1 or 2.")
                        continue

                    if Found:
                        print(f"Name '{name}' found between rows {start_row} and {end_row}")
                        last_recognized_face[name] = True  # Mark this face as already printed for this session

        # Display the video feed with rectangles and names
        cv2.imshow("Webcam", img)

        if cv2.waitKey(1) & 0xFF == ord('q'):  # Press 'q' to quit
            break
except KeyboardInterrupt:
    print("Program interrupted by user.")
except Exception as e:
    print(f"An error occurred: {e}")
finally:
    # Save the workbook after marking attendance
    workbook.save(file_path)
    print("Workbook saved successfully.")
    cap.release()
    cv2.destroyAllWindows()

    # Print the current working directory for debugging
print("Current working directory:", os.getcwd())
print("Attempting to open:", file_path)

# Load the workbook
try:
    workbook = load_workbook(file_path)
    # Perform your logic here...
    
    # After completing your processing, open the Excel file
    os.startfile(file_path)  # This will open the Excel file
    print("Excel file opened successfully.")

except FileNotFoundError as e:
    print("Error: File not found. Please check the file path.")
    print(e)

except Exception as e:
    print("An error occurred:", e)